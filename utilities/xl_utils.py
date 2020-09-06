import openpyxl


class xl:
    @staticmethod
    def get_xl_sheet_name(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        return workbook[sheetName]

    @staticmethod
    def getRowCount(file_name, sheetName):
        return xl.get_xl_sheet_name(file_name, sheetName).max_row

    @staticmethod
    def getColumnCount(file_name, sheetName):
        return xl.get_xl_sheet_name(file_name, sheetName).max_column

    @staticmethod
    def readData(file_name, sheetName, rownum, columnno):
        sheet = xl.get_xl_sheet_name(file_name, sheetName)
        return sheet.cell(row=rownum, column=columnno).value

    @staticmethod
    def writeData(file_name, sheetName, rownum, columnno, data):
        sheet = xl.get_xl_sheet_name(file_name, sheetName)
        sheet.cell(row=rownum, column=columnno).value = data
        openpyxl.workbook.save(file_name)
